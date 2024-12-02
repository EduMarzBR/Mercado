import pandas as pd
import glob
import os
from datetime import datetime
import warnings
import MetaTrader5 as mt5
import pandas_market_calendars as mcal
import talib
import numpy as np
import pandas_market_calendars as mcal
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Color

warnings.filterwarnings("ignore")

# Iniciar a conexão
if not mt5.initialize():
    print("Initialize() failed, error code =", mt5.last_error())
    quit()

# Verificar se a conexão foi bem-sucedida
print("MetaTrader5 version:", mt5.version())
print("")


# Lista de ações
# Permite que o usuário escolha entre o IBOV, IBRA, BDRX, ICON, IDIV, IEEX, IFIX, IFNC, IMAT, IMOB, INDX, MLCX, SMLL, UTIL

# Exibe a lista de códigos de índices
print("Escolha um índice:")
print("1 - IBOV")
print("2 - IBRA")
print("3 - BDRX")
print("4 - ICON")
print("5 - IDIV")
print("6 - IEEX")
print("7 - IFIX")
print("8 - IFNC")
print("9 - IMAT")
print("10 - IMOB")
print("11 - INDX")
print("12 - MLCX")
print("13 - SMLL")
print("14 - UTIL")

# Solicita ao usuário que digite o número correspondente ao índice desejado
codigo = int(input("Digite o número do índice desejado: "))

# Dicionário que mapeia códigos para os nomes dos índices
indices = {
    1: "IBOV",
    2: "IBRA",
    3: "BDRX",
    4: "ICON",
    5: "IDIV",
    6: "IEEX",
    7: "IFIX",
    8: "IFNC",
    9: "IMAT",
    10: "IMOB",
    11: "INDX",
    12: "MLCX",
    13: "SMLL",
    14: "UTIL",
}

# Verifica se o código digitado pelo usuário está no dicionário
if codigo in indices:
    indice = indices[codigo]
    print(f"Você selecionou o índice {indice}.")
else:
    print("Código de índice inválido.")

# Local onde está o arquivo CSV que contém a lista de ações do índice selecionado
list_of_files = glob.glob(
    f"C:/Users/armen/OneDrive/Estratégias/Listas/{indice}/*"
)  # * means all if need specific format then *.csv
latest_file = max(list_of_files, key=os.path.getctime)
# Carrega o arquivo CSV
lista = pd.read_csv(
    latest_file,
    sep=";",
    encoding="ISO-8859-1",
    skiprows=0,
    skipfooter=2,
    engine="python",
    thousands=".",
    decimal=",",
    header=1,
    index_col=False,
)


df = pd.DataFrame(lista)

acoes = df["Código"]

# Traz a data atual
data_atual = datetime.today()
dt = data_atual.strftime("%d/%m/%Y")

# Definir o intervalo de datas
inicio = "2010-01-01"
hoje = pd.Timestamp.now().strftime("%Y-%m-%d")


# Obter o calendário de negócios do Brasil
calendario = mcal.get_calendar(
    "B3"
)  # B3 é o código do calendário de negociação da Bolsa de Valores do Brasil

# Gerar datas úteis
dias_uteis = calendario.schedule(start_date=inicio, end_date=hoje)

# Contar o número de dias úteis
numero_dias_uteis = len(dias_uteis)


distante_media_movel = []
dias_ultimo_toque = []

for acao in acoes:
    df = mt5.copy_rates_from(
        acao, mt5.TIMEFRAME_D1, datetime.today(), numero_dias_uteis
    )

    df = pd.DataFrame(df)


    if len(df) == 0:
        print(f"Erro ao obter dados da ação {acao}.")
        continue


    df = pd.DataFrame(df)
    df["time"] = pd.to_datetime(df["time"], unit="s")
    df.set_index("time", inplace=True)

    close = df["close"].iloc[-1]
    mma200 = talib.MA(df["close"], timeperiod=200)

    mma200 = mma200.iloc[-1]

    dif_close_mma = mma200 - close
    d_media_movel = (dif_close_mma / close) * 100



    # Passo 1: Calcular a média móvel de 200 dias
    df["SMA_200"] = df["close"].rolling(window=200).mean()



    # Passo 2: Identificar quando o preço toca ou cruza a média móvel
    df["touch_or_cross"] = (df["close"] == df["SMA_200"]) | (
        (df["high"] >= df["SMA_200"]) & (df["low"] <= df["SMA_200"])
    )

    # Passo 3: Identificar os cruzamentos e toques
    touch_dates = df[df["touch_or_cross"]].index

    touch_dates_mean = touch_dates.diff().mean()
    touch_dates_mean = touch_dates_mean / np.timedelta64(1, "D")
    print(f"{acao}: --------------------------------------------------------")
    print(f"A distância percentual da média de 200 dias é de {d_media_movel:.2f}%")
    distante_media_movel.append(d_media_movel)

#    print(f"O tempo médio entre toques ou cruzamentos é de {touch_dates_mean:.2f} dias.")
    # Passo 4: Calcular o número de dias desde o último toque ou cruzamento
    if not touch_dates.empty:
        last_touch_date = touch_dates[-1]
        current_date = df.index[-1]
        # Crie um calendário de mercado para a B3
        calendario_b3 = mcal.get_calendar("BMF")

        days_since_last_touch = calendario_b3.valid_days(last_touch_date,current_date)
        days_since_last_touch = len(days_since_last_touch)
    else:
        days_since_last_touch = float("nan")

    print(
        f"O preço está distante da média móvel de 200 dias há {days_since_last_touch:.2f} dias desde o último toque ou cruzamento."
    )
    dias_ultimo_toque.append(days_since_last_touch)

df_dist = pd.DataFrame({"Código": acoes, "Dist. % da MMA 200": distante_media_movel, "Dias desde último toque": dias_ultimo_toque})
df_dist["Dist. % da MMA 200"] = df_dist["Dist. % da MMA 200"].apply(lambda x: round(x, 2))



total_acoes = len(df_dist)
print("")
print(f"Total de ações: {total_acoes}")

df_dist_na_media = df_dist[df_dist["Dist. % da MMA 200"] == 0].sort_values(by="Dist. % da MMA 200", ascending=False)
print("")
print(f"Quantidade de ações do ${indice} que estão na MMA 200: {len(df_dist_na_media)}")



df_dist_acima_10 = df_dist[df_dist["Dist. % da MMA 200"] < 0].sort_values(by="Dist. % da MMA 200", ascending=True).head(10)

df_dist_acima = df_dist[df_dist["Dist. % da MMA 200"] < 0].sort_values(by="Dist. % da MMA 200", ascending=True)
print("")
print(f"Quantidade de ações do ${indice} que estão acima da MMA 200: {len(df_dist_acima)}")

df_dist_abaixo_10 = df_dist[df_dist["Dist. % da MMA 200"] > 0].sort_values(by="Dist. % da MMA 200", ascending=False).head(10)
df_dist_abaixo = df_dist[df_dist["Dist. % da MMA 200"] > 0].sort_values(by="Dist. % da MMA 200", ascending=False)
print("")
print(f"Quantidade de ações do ${indice} que estão abaixo da MMA 200: {len(df_dist_abaixo)}")


print("")
print(f"TOP 10 ações mais distantes e abaixo da MMA 200 do ${indice}:")
print("")
print(df_dist_abaixo.head(10))

df_dist_abaixo_10.to_excel(rf"C:\Users\armen\OneDrive\Estratégias\Dist_Abaixo_MMA200_Top10_{indice}.xlsx", index=False)

# Carregar a planilha existente
wb = load_workbook(rf"C:\Users\armen\OneDrive\Estratégias\Dist_Abaixo_MMA200_Top10_{indice}.xlsx")
ws = wb.active  # Ou você pode selecionar uma folha específica, por exemplo: wb['NomeDaFolha']
# Descobrir a última linha na planilha
lr = ws.max_row



# Estilizar células A1:G1
for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=3):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
        cell.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
        cell.font = Font(color='FFA500',bold=True)


# Estilizar células A1:G1 (interior)
for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=3):
    for cell in row:
        cell.fill = PatternFill(patternType='solid', fgColor='000000')

# Definir cores alternadas
cor1 = '000000'  # Cor para linhas ímpares
cor2 = 'FFA500'  # Cor para linhas pares

# Iterar sobre as linhas e aplicar cores alternadas
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3), start=2):
    cor = cor1 if i % 2 == 1 else cor2
    for cell in row:
        cell.fill = PatternFill(start_color=cor, end_color=cor, fill_type='solid')

# Definir cores alternadas
cor1 = '00FFFFFF'  # Cor para linhas ímpares
cor2 = '000000'  # Cor para linhas pares

# Iterar sobre as linhas e aplicar cores alternadas
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3), start=2):
    cor = cor1 if i % 2 == 1 else cor2
    for cell in row:
        cell.font = Font(color = cor, bold=False)


# Salvar a planilha com os estilos aplicados
wb.save(rf"C:\Users\armen\OneDrive\Estratégias\Dist_Abaixo_MMA200_Top10_{indice}.xlsx")


print("")
print(f"TOP 10 ações mais distantes e acima da MMA 200 do ${indice}:")
print("")
print(df_dist_acima.head(10))

df_dist_acima_10.to_excel(rf"C:\Users\armen\OneDrive\Estratégias\Dist_Acima_MMA200_Top10_{indice}.xlsx", index=False)

# Carregar a planilha existente
wb = load_workbook(rf"C:\Users\armen\OneDrive\Estratégias\Dist_Acima_MMA200_Top10_{indice}.xlsx")
ws = wb.active  # Ou você pode selecionar uma folha específica, por exemplo: wb['NomeDaFolha']
# Descobrir a última linha na planilha
lr = ws.max_row



# Estilizar células A1:G1
for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=3):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
        cell.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
        cell.font = Font(color='FFA500',bold=True)


# Estilizar células A1:G1 (interior)
for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=3):
    for cell in row:
        cell.fill = PatternFill(patternType='solid', fgColor='000000')

# Definir cores alternadas
cor1 = '000000'  # Cor para linhas ímpares
cor2 = 'FFA500'  # Cor para linhas pares

# Iterar sobre as linhas e aplicar cores alternadas
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3), start=2):
    cor = cor1 if i % 2 == 1 else cor2
    for cell in row:
        cell.fill = PatternFill(start_color=cor, end_color=cor, fill_type='solid')

# Definir cores alternadas
cor1 = '00FFFFFF'  # Cor para linhas ímpares
cor2 = '000000'  # Cor para linhas pares

# Iterar sobre as linhas e aplicar cores alternadas
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3), start=2):
    cor = cor1 if i % 2 == 1 else cor2
    for cell in row:
        cell.font = Font(color = cor, bold=False)


# Salvar a planilha com os estilos aplicados
wb.save(rf"C:\Users\armen\OneDrive\Estratégias\Dist_Acima_MMA200_Top10_{indice}.xlsx")